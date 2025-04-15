# -*- coding: utf-8 -*-
import openpyxl
from pyrogram import Client
from datetime import datetime, timedelta

api_id = ''
api_hash = ''

# Задаем параметры непосредственно в коде
months_period = 3  # Количество месяцев для сохранения постов
channels = ["@hackproglib"]  # Список каналов для проверки


def process_channel(client, channel):
    print(f"\nОбработка канала: {channel}")

    # Создаем новую книгу Excel для каждого канала
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Заполняем заголовки столбцов
    sheet['A1'] = 'Канал'
    sheet['B1'] = 'Ссылка на пост'
    sheet['C1'] = 'Количество просмотров'
    sheet['D1'] = 'Количество репостов'
    sheet['E1'] = 'Сумма реакций'
    sheet['F1'] = 'Дата публикации'

    # Счетчик строк и столбцов
    row = 2
    column = 7  # Столбцы с эмодзи начинаются с седьмого столбца

    # Словарь для хранения всех возможных реакций
    all_reactions = {}

    # Вычисляем дату, начиная с которой нужно сохранять посты
    start_date = datetime.now() - timedelta(days=30 * months_period)

    try:
        # Получаем все сообщения
        messages = []
        for message in client.get_chat_history(channel):
            # Проверяем дату поста
            if message.date < start_date:
                print(f"Найден пост старше {months_period} месяца, переходим к следующему каналу")
                break

            # Добавляем сообщение в список для дальнейшей обработки
            messages.append(message)

        # Группируем сообщения по времени публикации (с точностью до 5 секунд)
        message_groups = {}

        for message in messages:
            # Получаем timestamp сообщения
            msg_time = int(message.date.timestamp())

            # Ищем группу для сообщения (проверяем разницу в 5 секунд)
            group_found = False
            for group_time in list(message_groups.keys()):
                if abs(group_time - msg_time) <= 5:
                    message_groups[group_time].append(message)
                    group_found = True
                    break

            # Если группа не найдена, создаем новую
            if not group_found:
                message_groups[msg_time] = [message]

        # Выбираем из каждой группы сообщение с наибольшим количеством просмотров
        filtered_messages = []

        for group_time, group_messages in message_groups.items():
            if len(group_messages) == 1:
                # Если в группе только одно сообщение, добавляем его
                filtered_messages.append(group_messages[0])
            else:
                # Если в группе несколько сообщений, находим с максимальным числом просмотров
                max_views_message = max(group_messages, key=lambda x: x.views if x.views is not None else 0)
                filtered_messages.append(max_views_message)
                print(
                    f"Группа из {len(group_messages)} сообщений: выбрано сообщение ID {max_views_message.id} с {max_views_message.views} просмотрами")

        # Сортируем сообщения по дате (от новых к старым)
        filtered_messages.sort(key=lambda x: x.date, reverse=True)

        # Обрабатываем отфильтрованные сообщения
        for message in filtered_messages:
            print('ID поста: ', message.id)
            try:
                # Проверяем количество просмотров
                views = message.views if message.views is not None else 0
                if views == 0:
                    print(f"Пропускаем пост {message.id} (нет просмотров)")
                    continue

                # Заполняем первый столбец названием канала
                sheet.cell(row=row, column=1, value=channel)

                # Заполняем второй столбец ссылкой на пост
                post_link = f"https://t.me/{channel[1:]}/{message.id}"
                sheet.cell(row=row, column=2).hyperlink = post_link
                sheet.cell(row=row, column=2).value = post_link
                sheet.cell(row=row, column=2).style = "Hyperlink"

                # Заполняем третий столбец количеством просмотров
                sheet.cell(row=row, column=3, value=views)

                # Заполняем четвертый столбец количеством репостов
                forwards = message.forwards if message.forwards is not None else 0
                sheet.cell(row=row, column=4, value=forwards)

                # Проверяем, есть ли реакции на сообщение
                total_reactions = 0
                if message.reactions:
                    for reaction in message.reactions.reactions:
                        emoji = reaction.emoji
                        count = reaction.count
                        total_reactions += count
                        if emoji not in all_reactions:
                            all_reactions[emoji] = column
                            sheet.cell(row=1, column=column, value=emoji)
                            column += 1
                        sheet.cell(row=row, column=all_reactions[emoji], value=count)

                # Записываем сумму реакций в пятый столбец
                sheet.cell(row=row, column=5, value=total_reactions)

                # Записываем дату публикации в шестой столбец
                date = message.date.strftime("%Y-%m-%d %H:%M:%S")
                sheet.cell(row=row, column=6, value=date)

                row += 1
            except Exception as e:
                print(f"Ошибка при обработке сообщения {message.id}: ", e)

        # Заполняем нулями отсутствующие реакции
        for r in range(2, row):
            for c in range(7, column):
                if sheet.cell(row=r, column=c).value is None:
                    sheet.cell(row=r, column=c, value=0)

        # Сохраняем книгу Excel
        xlsx_name = f"{channel[1:]}_telegram_stats_{months_period}months.xlsx"
        workbook.save(xlsx_name)
        print(f"Данные канала {channel} сохранены в файл: {xlsx_name}")

    except Exception as e:
        print(f"Ошибка при обработке канала {channel}: ", e)


# Основной код
with Client("my_account", api_id, api_hash) as client:
    for channel in channels:
        process_channel(client, channel)

print("\nОбработка всех каналов завершена")
