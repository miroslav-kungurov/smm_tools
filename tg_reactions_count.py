# -*- coding: utf-8 -*-
import openpyxl
from pyrogram import Client

api_id = ''
api_hash = ''

# Создаем новую книгу Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

# Заполняем заголовки столбцов
sheet['A1'] = 'Канал'
sheet['B1'] = 'Ссылка на пост'
sheet['C1'] = 'Количество просмотров'
sheet['D1'] = 'Сумма реакций'

# Счетчик строк и столбцов
row = 2
column = 5  # Столбцы с эмодзи начинаются с пятого столбца

# Словарь для хранения всех возможных реакций
all_reactions = {}

first_post_id = 1
last_post_id = 9248
channel = "@progbook"

with Client("my_account", api_id, api_hash) as client:
    for message_id in range(first_post_id, last_post_id):
        print('ID поста: ', message_id)
        try:
            message = client.get_messages(chat_id=channel, message_ids=message_id)

            # Проверяем количество просмотров
            views = message.views if message.views is not None else 0
            if views == 0:
                print(f"Пропускаем пост {message_id} (нет просмотров)")
                continue  # Пропускаем этот пост и переходим к следующему

            # Заполняем первый столбец названием канала
            sheet.cell(row=row, column=1, value=channel)

            # Заполняем второй столбец ссылкой на пост
            post_link = f"https://t.me/{channel[1:]}/{message_id}"
            sheet.cell(row=row, column=2).hyperlink = post_link
            sheet.cell(row=row, column=2).value = post_link
            sheet.cell(row=row, column=2).style = "Hyperlink"

            # Заполняем третий столбец количеством просмотров
            sheet.cell(row=row, column=3, value=views)

            # Проверяем, есть ли реакции на сообщение
            total_reactions = 0
            if message.reactions:
                for reaction in message.reactions.reactions:
                    emoji = reaction.emoji
                    count = reaction.count
                    total_reactions += count
                    if emoji not in all_reactions:
                        all_reactions[emoji] = column
                        sheet.cell(row=1, column=column, value=emoji)
                        column += 1
                    sheet.cell(row=row, column=all_reactions[emoji], value=count)

            # Записываем сумму реакций в четвертый столбец
            sheet.cell(row=row, column=4, value=total_reactions)

            row += 1
        except Exception as e:
            print(f"Error getting message {message_id}: ", e)

# Заполняем нулями отсутствующие реакции
for r in range(2, row):
    for c in range(5, column):
        if sheet.cell(row=r, column=c).value is None:
            sheet.cell(row=r, column=c, value=0)

# Сохраняем книгу Excel
xlsx_name = channel[1:] + "_telegram_reactions.xlsx"
workbook.save(xlsx_name)
